from io import BytesIO
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type="solid", fgColor="A51C30")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="D7DEE8"),
    right=Side(style="thin", color="D7DEE8"),
    top=Side(style="thin", color="D7DEE8"),
    bottom=Side(style="thin", color="D7DEE8"),
)


def _autosize_columns(ws, rows: Sequence[Sequence[object]]) -> None:
    if not rows:
        return

    widths: list[int] = [0] * len(rows[0])
    for row in rows:
        for idx, value in enumerate(row):
            display_value = "" if value is None else str(value)
            widths[idx] = max(widths[idx], len(display_value))

    for idx, width in enumerate(widths, start=1):
        column = get_column_letter(idx)
        ws.column_dimensions[column].width = max(12, min(width + 2, 42))


def _render_sheet(ws, headers: Sequence[str], rows: Iterable[Sequence[object]]) -> None:
    ws.append(list(headers))
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    frozen_rows = [headers]
    for row in rows:
        ws.append(list(row))
        frozen_rows.append(list(row))

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    _autosize_columns(ws, frozen_rows)


def _workbook_bytes(ws_title: str, headers: Sequence[str], rows: Iterable[Sequence[object]]) -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = ws_title
    _render_sheet(ws, headers, rows)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()


def build_class_students_export(class_item, student_items) -> bytes:
    headers = [
        "Mã sinh viên",
        "Họ tên",
        "Email",
        "Số điện thoại",
        "Lớp hành chính",
        "Có mặt",
        "Đi muộn",
        "Vắng mặt",
        "Đã nộp bài",
        "Thiếu bài",
    ]
    rows = [
        [
            item.get("student_code"),
            item.get("full_name"),
            item.get("email"),
            item.get("phone"),
            item.get("class_name"),
            item.get("present_count", 0),
            item.get("late_count", 0),
            item.get("absent_count", 0),
            item.get("submitted_count", 0),
            item.get("missing_count", 0),
        ]
        for item in student_items
    ]
    return _workbook_bytes(f"Lop_{class_item['class_code']}", headers, rows)


def build_attendance_session_export(session_item, attendance_items) -> bytes:
    headers = [
        "Mã sinh viên",
        "Họ tên",
        "Email",
        "Trạng thái điểm danh",
        "Thời gian check-in",
        "Ghi chú",
    ]
    rows = [
        [
            item.get("student_code"),
            item.get("full_name"),
            item.get("email"),
            item.get("attendance_status"),
            item.get("checkin_time"),
            item.get("note"),
        ]
        for item in attendance_items
    ]
    return _workbook_bytes(f"Buoi_{session_item['id']}", headers, rows)


def build_assignment_scores_export(assignment_item, submission_items) -> bytes:
    headers = [
        "Mã sinh viên",
        "Họ tên",
        "Email",
        "Trạng thái nộp bài",
        "Thời gian nộp",
        "Điểm",
        "Nhận xét",
    ]
    rows = [
        [
            item.get("student_code"),
            item.get("full_name"),
            item.get("email"),
            item.get("status") or "missing",
            item.get("submitted_at"),
            item.get("teacher_score"),
            item.get("teacher_comment"),
        ]
        for item in submission_items
    ]
    return _workbook_bytes(f"BaiTap_{assignment_item['id']}", headers, rows)
